import json
import re
import sys

from PyQt6.QtWidgets import QMainWindow, QGridLayout, QPushButton, QWidget, QScrollArea, QApplication, QVBoxLayout, \
    QCheckBox, QFileDialog
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
import pandas as pd
import os
from openpyxl.styles import PatternFill


class ApplicationWindowManager:
    def __init__(self):
        self.app = QApplication(sys.argv)
        self.current_window = None

    def start(self):
        self.show_main_menu()
        sys.exit(self.app.exec())

    def show_main_menu(self):
        if self.current_window is not None:
            self.current_window.close()

        self.current_window = ConvertPartsDataToExcel()


class ConvertPartsDataToExcel(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Import OrderablePart Raw")
        self.resize(950, 800)

        # Initialize the folder_path to None
        self.folder_path = None

        # Naming
        self.naming_section_layout = QVBoxLayout()
        self.choose_folder_btn = QPushButton("Ordner Auswählen", self)
        self.choose_folder_btn.clicked.connect(self.open_folder_dialog)
        self.same_folder_checkbox = QCheckBox("In gleichem Ordner speichern", self)
        self.convert_to_excel_btn = QPushButton("Alle Konvertieren", self)
        self.convert_to_excel_btn.clicked.connect(self.convert_to_excel)
        self.naming_section_layout.addWidget(self.choose_folder_btn)
        self.naming_section_layout.addWidget(self.same_folder_checkbox)
        self.naming_section_layout.addWidget(self.convert_to_excel_btn)

        # ------------------ setup Window ------------------
        central_widget = QWidget()
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setWidget(central_widget)

        screen = QApplication.primaryScreen()
        rect = screen.availableGeometry()
        self.setMaximumHeight(rect.height())

        self.setCentralWidget(scroll_area)
        self.main_layout = QVBoxLayout(central_widget)

        # ------------------ setup Sections ------------------

        self.main_layout.addLayout(self.naming_section_layout)

        self.show()

    def open_folder_dialog(self):
        file_dialog = QFileDialog()
        json_filter = "JSON (*.json)"
        folder = file_dialog.getExistingDirectory()
        # folder = QFileDialog.getExistingDirectory(self, "Select Directory", "", options=options)
        if folder:
            self.folder_path = folder

    def convert_to_excel(self):
        if self.folder_path:
            create_custom_formatted_excel(self.folder_path)
        else:
            # You can show an error message or notification here if folder_path is None.
            pass


def sanitize_content(content):
    """Sanitize content to ensure it's encodable in 'latin-1'."""
    return content.encode('latin-1', 'ignore').decode('latin-1')


def sanitize_filename(filename):
    # Replace non-ASCII characters with their closest ASCII equivalent
    filename = filename.encode('ascii', 'ignore').decode()

    # Replace invalid filename characters with "_"
    filename = re.sub(r'[\\/*?:"<>|]', '_', filename)

    # Limit the length of the filename to 200 characters
    if len(filename) > 200:
        filename = filename[:200]

    return filename


def create_custom_formatted_excel(directory_path):
    # List all files in the directory
    files = [f for f in os.listdir(directory_path) if
             os.path.isfile(os.path.join(directory_path, f)) and f.endswith(".json")]

    for file in files:
        # Read the JSON content
        with open(os.path.join(directory_path, file), "r", encoding="utf-8") as f:
            data = json.load(f)

        # Extract required information and put it into a DataFrame
        df_data = []
        for entry in data:
            index = sanitize_content(entry['values'].get('pos', ''))
            teilenummer = sanitize_content(entry['values'].get('partno', ''))
            name = sanitize_content(entry.get('description', '').replace('\n', ' '))
            kommentar = sanitize_content((entry['values'].get('comment', '').replace('\n', ' ')))
            menge = sanitize_content(entry['values'].get('qty', ''))
            df_data.append([index, teilenummer, name, kommentar, menge])

        # Create a new Excel workbook and worksheet
        wb = Workbook()
        ws = wb.active

        # Apply desired formatting
        font = Font(name='Arial', size=12)
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
        alignment = Alignment(wrap_text=True)

        # Hardcode column widths
        col_widths = [4, 10, 28, 30, 4]
        for col_index, col_width in enumerate(col_widths, start=1):
            ws.column_dimensions[chr(64 + col_index)].width = col_width

        # Extracting the heading from the filename
        filename_parts = file.replace(" - ", "_").split("_")
        heading_parts = filename_parts[5:]
        heading_parts[-1] = heading_parts[-1].replace('.json', '')
        heading = sanitize_content(" - ".join(heading_parts))

        # Insert merged cells for heading
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        ws['A1'].value = heading
        ws['A1'].font = font
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # Set header background color
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        headers = ["Position", "Teilenummer", "Name", "Kommentar", "Menge"]
        for col_index, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_index, value=header)
            cell.font = font
            cell.border = border
            cell.alignment = alignment
            cell.fill = header_fill

        # Write data to Excel worksheet
        row_height = 40
        for row_index, row in enumerate(df_data, start=2):
            for col_index, cell_value in enumerate(row, start=1):
                cell = ws.cell(row=row_index, column=col_index, value=cell_value)
                cell.font = font
                cell.border = border
                cell.alignment = alignment
                if col_index in [3, 4]:  # Name and Kommentar columns
                    if col_index is 3:
                        lines_required = -(-len(cell_value) // 20)
                    else:
                        lines_required = -(-len(cell_value) // 30)
                    calc_row_height = 40 + (max(0, lines_required - 2) * 15)
                    row_height = calc_row_height if calc_row_height > row_height else row_height
                    ws.row_dimensions[row_index].height = row_height
                if col_index is 5:
                    row_height = 40


        # Sanitize the heading for use as a filename
        excel_filename = sanitize_filename(heading) + ".xlsx"

        # Save the Excel workbook
        wb.save(os.path.join(directory_path, excel_filename))

    return f"Processed {len(files)} files."


if __name__ == '__main__':
    manager = ApplicationWindowManager()
    manager.start()