import json
import re
import sys
from PIL import Image
import PyPDF2
from PyPDF2 import PdfWriter, PdfReader

from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import QMainWindow, QGridLayout, QPushButton, QWidget, QScrollArea, QApplication, QVBoxLayout, \
    QCheckBox, QFileDialog, QLabel, QMessageBox
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
import pandas as pd
import os
from openpyxl.styles import PatternFill
import win32com.client as win32

from core.UI.NavigationBar import NavigationBar
from dotenv import load_dotenv
load_dotenv()
DEFAULT_PARTS_JSON_FOLDER_PATH = os.getenv("DEFAULT_PARTS_JSON_FOLDER_PATH")
A4_DIMENSIONS = (2480, 3508)  # At 300 DPI
A4_DPI = (300, 300)  # Standard printing DPI


class ConvertPartsToExcelList(QMainWindow):
    def __init__(self, window_manager):
        super().__init__()
        self.setWindowTitle("Convert Parts to Excel List")
        self.resize(600, 300)

        # Initialize the folder_path to None
        self.folder_path = None

        # Naming
        self.naming_section_layout = QVBoxLayout()
        self.folder_path_label = QLabel(DEFAULT_PARTS_JSON_FOLDER_PATH, self)
        self.choose_folder_btn = QPushButton("Ordner Auswählen", self)
        self.convert_to_excel_btn = QPushButton("Alle Konvertieren", self)
        self.combine_exports_btn = QPushButton("Alle Exporte kombinieren", self)

        self.naming_section_layout.addWidget(self.folder_path_label)
        self.naming_section_layout.addWidget(self.choose_folder_btn)
        self.naming_section_layout.addWidget(self.convert_to_excel_btn)
        self.naming_section_layout.addWidget(self.combine_exports_btn)

        self.choose_folder_btn.clicked.connect(self.open_folder_dialog)
        self.convert_to_excel_btn.clicked.connect(self.convert_to_excel)
        self.combine_exports_btn.clicked.connect(self.merge_pdfs)

        # ------------------ setup Window ------------------
        central_widget = QWidget()
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setWidget(central_widget)

        screen = QApplication.primaryScreen()
        rect = screen.availableGeometry()
        self.setMaximumHeight(rect.height())

        self.setCentralWidget(scroll_area)
        self.main_layout = QVBoxLayout(central_widget)

        # Window Manager
        self.window_manager = window_manager
        self.nav_bar = NavigationBar(self.window_manager)
        self.main_layout.addWidget(self.nav_bar)
        self.main_layout.setAlignment(self.nav_bar, Qt.AlignmentFlag.AlignTop)

        # ------------------ setup Sections ------------------

        self.main_layout.addLayout(self.naming_section_layout)

        self.show()

    def open_folder_dialog(self):
        file_dialog = QFileDialog()
        folder = file_dialog.getExistingDirectory(directory=r"C:\Users\TheOverlanders\Downloads\parts_link_exports")
        # folder = QFileDialog.getExistingDirectory(self, "Select Directory", "", options=options)
        if folder:
            self.folder_path = folder
            self.folder_path_label.text = folder

    def convert_to_excel(self):
        try:
            if self.folder_path:
                convert_json_to_excel_and_pdf(self.folder_path)
                message = convert_json_to_excel_and_pdf(self.folder_path)
                QMessageBox.information(self, "Erfolgreich", message)
            else:
                self.folder_path = DEFAULT_PARTS_JSON_FOLDER_PATH
                message = convert_json_to_excel_and_pdf(self.folder_path)
                QMessageBox.information(self, "Erfolgreich", message)
        except Exception as e:
            print(e.__str__())
            QMessageBox.warning(self, "Error", "While parsing Data: " + e.__str__())

    def merge_pdfs(self, output_filename="combined.pdf"):
        try:
            if self.folder_path:
                convert_json_to_excel_and_pdf(self.folder_path)
                message = merge(self.folder_path)
                QMessageBox.information(self, "Erfolgreich", message)
            else:
                self.folder_path = DEFAULT_PARTS_JSON_FOLDER_PATH
                message = merge(self.folder_path)
                QMessageBox.information(self, "Erfolgreich", message)
        except Exception as e:
            print(e.__str__())
            QMessageBox.warning(self, "Error", "While parsing Data: " + e.__str__())

def sanitize_content(content):
    """Sanitize content to ensure it's encodable in 'latin-1'."""
    return content.encode('latin-1', 'ignore').decode('latin-1')


def sanitize_filename(filename):
    # Replace non-ASCII characters with their closest ASCII equivalent
    filename = filename.encode('ascii', 'ignore').decode()

    # Replace invalid filename characters with "_"
    filename = re.sub(r'[\\/*?:"<>|]', '_', filename)

    # Limit the length of the filename to 200 characters
    if len(filename) > 200:
        filename = filename[:200]

    return filename


def convert_json_to_excel_and_pdf(directory_path):
    # List all files in the directory and its subdirectories
    iter = 0
    json_files = []
    png_files = []
    for dirpath, dirnames, filenames in os.walk(directory_path):
        for filename in [f for f in filenames]:
            if filename.endswith(".json"):
                json_files.append(os.path.join(dirpath, filename))
            elif filename.endswith(".png"):
                png_files.append(os.path.join(dirpath, filename))

    # make jsons to pdfs
    for json_file in json_files:
        with open(json_file, "r", encoding="utf-8") as f:
            data = json.load(f)

        # Extract required information and put it into a DataFrame
        df_data = []
        for entry in data:
            index = sanitize_content(entry['values'].get('pos', ''))
            teilenummer = sanitize_content(entry['values'].get('partno', ''))
            name = sanitize_content(entry.get('description', '').replace('\n', ' '))
            kommentar = sanitize_content((entry['values'].get('comment', '').replace('\n', ' ')))
            menge = sanitize_content(entry['values'].get('qty', ''))
            df_data.append([index, teilenummer, name, kommentar, menge])

        # Create a new Excel workbook and worksheet
        wb = Workbook()
        ws = wb.active

        # Apply desired formatting
        font = Font(name='Arial', size=12)
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
        alignment = Alignment(wrap_text=True)

        # Hardcode column widths
        col_widths = [4, 10, 28, 30, 4]
        for col_index, col_width in enumerate(col_widths, start=1):
            ws.column_dimensions[chr(64 + col_index)].width = col_width

        # Extracting the heading from the filename
        try:
            filename_parts = json_file.replace(" - ", "_").split("_")
            heading_parts = filename_parts[5:]
            heading_parts[-1] = heading_parts[-1].replace('.json', '')
            heading = sanitize_content(" - ".join(heading_parts))
        except Exception as e:
            heading = json_file.replace('.json', '')

        # Insert merged cells for heading
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        ws['A1'].value = heading
        ws['A1'].font = font
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Set header background color
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        headers = ["Position", "Teilenummer", "Name", "Kommentar", "Menge"]
        for col_index, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_index, value=header)
            cell.font = font
            cell.border = border
            cell.alignment = alignment
            cell.fill = header_fill

        # Write data to Excel worksheet
        row_height = 40
        for row_index, row in enumerate(df_data, start=3):
            for col_index, cell_value in enumerate(row, start=1):
                cell = ws.cell(row=row_index, column=col_index, value=cell_value)
                cell.font = font
                cell.border = border
                cell.alignment = alignment
                if col_index in [3, 4]:  # Name and Kommentar columns
                    if col_index is 3:
                        lines_required = -(-len(cell_value) // 20)
                    else:
                        lines_required = -(-len(cell_value) // 30)
                    calc_row_height = 40 + (max(0, lines_required - 2) * 15)
                    row_height = calc_row_height if calc_row_height > row_height else row_height
                    ws.row_dimensions[row_index].height = row_height
                if col_index is 5:
                    row_height = 40

        # Sanitize the heading for use as a filename
        excel_filename = json_file.replace('.json','.xlsx')

        # Save the Excel workbook
        wb.save(excel_filename)

        # Convert the Excel workbook to PDF
        # parent_dir = os.path.dirname(directory_path.rstrip(os.sep))
        # pdf_filename = directory_path + '/' + os.path.basename(file).replace('.json','.pdf')
        pdf_filename = directory_path + '\\' + "export_" + str(iter) + ".1" + ".pdf"
        excel_to_pdf(excel_filename,  pdf_filename)

        rescale_pdf(pdf_filename, pdf_filename, 2)
        iter += 1

    print(f"Converted {len(json_files)} JSON files.")
    iter = 0
    for png_file in png_files:
        with Image.open(png_file) as img:
            # Resize and pad image to A4 dimensions
            a4_img = resize_and_pad(img, A4_DIMENSIONS)

            # Convert PNG to PDF with specified DPI
            pdf_file = os.path.join(os.path.dirname(os.path.dirname(png_file)), f'export_{iter}.0.pdf')
            a4_img.save(pdf_file, dpi=A4_DPI)
            iter += 1
    print(f"Converted {len(png_files)} PNG files.")
    return f"Converted {len(json_files)} JSON files and {len(png_files)} PNG files."


def excel_to_pdf(excel_filename, pdf_filename):
    print("----------------------------------------")
    print("PDF:")
    print(pdf_filename)
    print("EXCEL:")
    print(excel_filename)
    excel = win32.Dispatch('Excel.Application')
    excel.Visible = False

    # Open the Excel workbook
    workbook = excel.Workbooks.Open(excel_filename)

    try:
        workbook.ExportAsFixedFormat(0, pdf_filename)
    except Exception as e:
        print(f"Failed to convert {excel_filename} to PDF. {str(e)}")
        raise e
    finally:
        workbook.Close()
        excel.Quit()


def resize_and_pad(img, target_dimensions):
    # Calculate aspect ratio of the original image and the target size
    src_aspect = img.width / img.height
    tgt_aspect = target_dimensions[0] / target_dimensions[1]

    # Resize the image
    if src_aspect > tgt_aspect:
        # If the original image is wider than the target size
        resized_img = img.resize((target_dimensions[0], round(target_dimensions[0] / src_aspect)))
    else:
        # If the original image is taller than the target size or equal
        resized_img = img.resize((round(target_dimensions[1] * src_aspect), target_dimensions[1]))

    # Create a white background
    output = Image.new("RGB", target_dimensions, "white")

    # Paste the resized image onto the white background
    y_offset = (target_dimensions[1] - resized_img.height) // 2
    x_offset = (target_dimensions[0] - resized_img.width) // 2
    output.paste(resized_img, (x_offset, y_offset))

    return output


def rescale_pdf(input_pdf, output_pdf, scale_factor):
    # Read the input
    reader = PdfReader(input_pdf)
    writer = PdfWriter()

    # Loop through all pages in the input PDF and scale each page
    for page in reader.pages:
        page.scale_by(scale_factor)
        writer.add_page(page)

    # Write the scaled pages to the output PDF
    writer.write(output_pdf)

def merge(directory_path, output_filename="combined.pdf"):
    # List all files in the directory
    files = [f for f in os.listdir(directory_path) if f.endswith('.pdf')]

    # Sort the files by name
    sorted_files = sorted(files)

    # Create a PDF merger object
    pdf_merger = PyPDF2.PdfMerger()

    # Append each PDF to the merger object
    for file in sorted_files:
        with open(os.path.join(directory_path, file), 'rb') as f:
            pdf_merger.append(f)

    # Write the merged PDF to the output file
    with open(os.path.join(directory_path, output_filename), 'wb') as f:
        pdf_merger.write(f)

    return f"PDFs merged into {output_filename}"