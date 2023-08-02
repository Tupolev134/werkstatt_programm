import re

from core.Profile import Profile
import pandas as pd
from openpyxl import load_workbook
from copy import copy

class ProfileController():

    def __init__(self, profile):
        self.profile = profile
        self.wb = load_workbook(self.profile.path_to_excel)
        self.wb_values = load_workbook(self.profile.path_to_excel, data_only=True)
        self.ws = self.wb.active
        self.ws_values = self.wb_values.active
        self.df = pd.read_excel(self.profile.path_to_excel)

    def populate_profile(self):
        # Reading the excel file with pandas
        df = pd.read_excel(self.profile.path_to_excel)

        # Looping through the DataFrame
        for index, row in df.iterrows():
            kategorie = row[0] if pd.notnull(row[0]) else "unidentified"
            artikelnummer = row[1] if pd.notnull(row[1]) else "unidentified"
            beschreibung = row[2] if pd.notnull(row[2]) else "unidentified"
            haendler = row[6] if pd.notnull(row[6]) else "unidentified"

            self.profile.add_kategorie(kategorie)
            self.profile.add_artikelnummer(artikelnummer)
            self.profile.add_beschreibung(beschreibung)
            self.profile.add_haendler(haendler)
        self.profile.save_to_json(self.profile.path_to_profile)

    def add_part_to_excel(self, kategorie, artikelnummer, beschreibung, menge : int, einzelpreis: float,gesamtpreis: float, haendler, darlene, diana, menge_da, menge_di):
        anteil_da = einzelpreis * menge_da
        anteil_di = einzelpreis * menge_di

        position = self.ws.max_row - 1
        self.insert_row_and_update_formulas(position)

        # Add data to the new row
        new_row = [kategorie,artikelnummer,beschreibung,menge,einzelpreis,gesamtpreis,haendler, darlene, diana, menge_da, menge_di, anteil_da, anteil_di]
        for col, data in enumerate(new_row, start=1):
            self.ws.cell(row=position, column=col, value=data)
        self.wb.save(self.profile.path_to_excel)

    def insert_row_and_update_formulas(self, position):
        print("position: " + str(position))

        # L
        di_formula = self.ws[f'M{position}'].value
        di_new_formula = re.sub(r'(\d+)(\))$', str(position) + r'\2', di_formula)
        self.ws[f'M{position}'].value = di_new_formula

        # M
        da_formula = self.ws[f'L{position}'].value
        da_new_formula = re.sub(r'(\d+)(\))$', str(position) + r'\2', da_formula)
        self.ws[f'L{position}'].value = da_new_formula

        # E
        gesamt_einzelpreis = self.ws[f'E{position+1}'].value
        gesamt_einzelpreis = re.sub(r'(\d+)(\))$', str(position) + r'\2', gesamt_einzelpreis)
        print(gesamt_einzelpreis)
        self.ws[f'E{position+1}'].value = gesamt_einzelpreis

        # F
        gesamt_gesamtpreis = self.ws[f'F{position+1}'].value
        gesamt_gesamtpreis = re.sub(r'(\d+)(\))$', str(position) + r'\2', gesamt_gesamtpreis)
        self.ws[f'F{position+1}'].value = gesamt_gesamtpreis

        total_right = self.ws[f'M{position+1}'].value
        def increment(match):
            return match.group(1) + str(int(match.group(2)) + 1)
        new_total_right = re.sub(r'(L|M)(\d+)', increment, total_right)
        self.ws[f'M{position + 1}'].value = new_total_right

        # Insert the new row
        self.ws.insert_rows(position)
        self.ws.row_dimensions[position].height = 40
        for col_num in range(1, self.ws.max_column + 1):
            source_cell = self.ws.cell(row=self.ws.max_row, column=col_num)
            target_cell = self.ws.cell(row=position, column=col_num)

            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)

            if source_cell.data_type == 'f':
                target_cell.value = source_cell.value
            else:
                target_cell.value = source_cell.value

    def get_gesamtpreis(self):
        position = self.ws.max_row - 1
        return self.ws_values[f'F{position+1}'].value

    def get_letzte_zeile(self):
        return self.ws.max_row - 1

if __name__ == '__main__':
    p = ProfileController()
    # p.populate_profile()
    p.add_part_to_excel("test_kat","test_art","test_besch",1,200000.43,4,"test_hae","DI","",1,0)
    p.add_part_to_excel("test_kat","test_art","test_besch",1,400000.43,4,"test_hae","DA","",0,1)
    p.add_part_to_excel("test_kat","test_art","test_besch",2,400000.43,4,"test_hae","DA","DI",1,1)